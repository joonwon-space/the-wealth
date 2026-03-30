"""포트폴리오 CSV/Excel 내보내기 API."""

import csv
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.logging import get_logger
from app.db.session import get_db
from app.models.holding import Holding
from app.models.portfolio import Portfolio
from app.models.transaction import Transaction
from app.models.user import User

router = APIRouter(prefix="/portfolios", tags=["portfolio-export"])
logger = get_logger(__name__)


def _assert_portfolio_owner(portfolio: Portfolio, user: User) -> None:
    if portfolio.user_id != user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized"
        )


@router.get("/{portfolio_id}/export/csv")
async def export_holdings_csv(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """보유 종목 CSV 내보내기."""
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )
    _assert_portfolio_owner(portfolio, current_user)

    result = await db.execute(
        select(Holding)
        .where(Holding.portfolio_id == portfolio_id)
        .order_by(Holding.ticker)
    )
    holdings = list(result.scalars().all())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ticker", "name", "quantity", "avg_price", "invested", "created_at"])
    for h in holdings:
        invested = h.quantity * h.avg_price
        writer.writerow([
            h.ticker,
            h.name,
            str(h.quantity),
            str(h.avg_price),
            str(invested),
            h.created_at.isoformat(),
        ])

    output.seek(0)
    filename = f"holdings_portfolio_{portfolio_id}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{portfolio_id}/transactions/export/csv")
async def export_transactions_csv(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """거래 내역 CSV 내보내기."""
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )
    _assert_portfolio_owner(portfolio, current_user)

    result = await db.execute(
        select(Transaction)
        .where(
            Transaction.portfolio_id == portfolio_id,
            Transaction.deleted_at.is_(None),
        )
        .order_by(Transaction.traded_at.desc())
    )
    transactions = list(result.scalars().all())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "ticker", "type", "quantity", "price", "total", "traded_at"])
    for txn in transactions:
        total = txn.quantity * txn.price
        writer.writerow([
            txn.id,
            txn.ticker,
            txn.type,
            str(txn.quantity),
            str(txn.price),
            str(total),
            txn.traded_at.isoformat(),
        ])

    output.seek(0)
    filename = f"transactions_portfolio_{portfolio_id}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _apply_header_style(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, columns: int) -> None:
    """헤더 행 스타일 적용: 볼드, 배경색."""
    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """열 너비 자동 조정 (최소 10, 최대 40)."""
    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        adjusted = min(max(max_length + 2, 10), 40)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted


@router.get("/{portfolio_id}/export/xlsx")
async def export_xlsx(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """포트폴리오 Excel(xlsx) 내보내기.

    시트 1: 보유 종목, 시트 2: 거래 내역
    """
    portfolio = await db.get(Portfolio, portfolio_id)
    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )
    _assert_portfolio_owner(portfolio, current_user)

    holdings_result = await db.execute(
        select(Holding)
        .where(Holding.portfolio_id == portfolio_id)
        .order_by(Holding.ticker)
    )
    holdings = list(holdings_result.scalars().all())

    txn_result = await db.execute(
        select(Transaction)
        .where(
            Transaction.portfolio_id == portfolio_id,
            Transaction.deleted_at.is_(None),
        )
        .order_by(Transaction.traded_at.desc())
    )
    transactions = list(txn_result.scalars().all())

    wb = openpyxl.Workbook()

    # --- 시트 1: 보유 종목 ---
    ws_holdings = wb.active
    ws_holdings.title = "보유종목"
    holding_headers = ["티커", "종목명", "수량", "평균단가", "투자금액", "등록일시"]
    ws_holdings.append(holding_headers)
    _apply_header_style(ws_holdings, 1, len(holding_headers))

    num_fmt = "#,##0.####"
    for h in holdings:
        invested = float(h.quantity) * float(h.avg_price)
        created_at_str = h.created_at.isoformat() if h.created_at else ""
        ws_holdings.append([
            h.ticker,
            h.name,
            float(h.quantity),
            float(h.avg_price),
            round(invested, 2),
            created_at_str,
        ])
        # 숫자 열 서식
        for col_idx in (3, 4, 5):
            ws_holdings.cell(row=ws_holdings.max_row, column=col_idx).number_format = num_fmt

    _auto_fit_columns(ws_holdings)

    # --- 시트 2: 거래 내역 ---
    ws_txns = wb.create_sheet("거래내역")
    txn_headers = ["ID", "티커", "유형", "수량", "단가", "총금액", "메모", "거래일시"]
    ws_txns.append(txn_headers)
    _apply_header_style(ws_txns, 1, len(txn_headers))

    for txn in transactions:
        total = float(txn.quantity) * float(txn.price)
        row_data = [
            txn.id,
            txn.ticker,
            txn.type,
            float(txn.quantity),
            float(txn.price),
            round(total, 2),
            txn.memo or "",
            txn.traded_at.isoformat() if txn.traded_at else "",
        ]
        ws_txns.append(row_data)
        for col_idx in (4, 5, 6):
            ws_txns.cell(row=ws_txns.max_row, column=col_idx).number_format = num_fmt

    _auto_fit_columns(ws_txns)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"portfolio_{portfolio_id}_{today}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
